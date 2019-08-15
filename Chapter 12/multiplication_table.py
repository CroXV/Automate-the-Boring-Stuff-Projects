#! /usr/bin/env python
# Usage: py.exe multiplication_table.py <number>

from openpyxl.styles import Font
import openpyxl
import sys


if len(sys.argv) == 2:
    multiply_num = int(sys.argv[1]) + 2
else:
    multiply_num = int(input('Multiplication Number:\n> ')) + 2

# Create new workbook
wb = openpyxl.Workbook()
ws = wb.active
bold = Font(bold=True)

for num in range(2, multiply_num):
    # first row without a1 formatted to bold
    ws.cell(row=1, column=num, value=num - 1).font = bold
    # first column without a1 formatted to bold
    ws.cell(row=num, column=1, value=num - 1).font = bold
    for col in range(2, multiply_num):
        ws.cell(row=num, column=col, value=(num - 1) * (col - 1))

wb.save('multiplication_table.xlsx')
print('Done.')
