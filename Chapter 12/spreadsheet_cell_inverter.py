#! /usr/bin/env python
# Usage: py.exe spreadsheet_cell_inverter.py <workbook>

import openpyxl
import sys


if len(sys.argv) == 2:
    file = sys.argv[1]
else:
    file = input('Enter Workbook path:\n> ')


wb = openpyxl.load_workbook(file)       # load file
ws = wb.active                          # get first worksheet
ws2 = wb.create_sheet()                 # create empty worksheet


for row in range(1, ws.max_row + 1):
    for col in range(1, ws.max_column + 1):
        # copy inverted rows and columns to new empty worksheet
        ws2.cell(row=col, column=row,
                 value=ws.cell(row=row, column=col).value)


wb.remove(ws)           # delete old worksheet
ws2.title = ws.title    # rename new worksheet with old worksheet's title

wb.save(file)           # save file
